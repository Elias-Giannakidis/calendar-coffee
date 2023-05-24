
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles.colors import Color
from pathManager import getNextMonthPath, getNextMonth, getPreviousMonth, getWeekDay
import datetime
import peopleManager


dayColor = PatternFill(fill_type='solid', start_color="EDD79F")
weekendColor = PatternFill(fill_type='solid', start_color="71EA8F")
color1 = PatternFill(fill_type='solid', start_color="DED6BA")
color2 = PatternFill(fill_type='solid', start_color="EFECE0")
line = Side(border_style='dashed', color='000000')
border = Border(top=line, bottom=line, right=line, left=line)

path = getNextMonthPath()

def genWorkSheet(wb, sheetName, month):
    thisMonth = datetime.date.today().month
    nextMonth = getNextMonth(thisMonth)
    prevMonth = getPreviousMonth(thisMonth)
    ws = wb.create_sheet(sheetName)
    daysAfter = 0
    if month == 'last':
        daysAfter = -64
    if month == 'this':
        daysAfter = -32
    run = True
    col = 1
    ws["A2"] = "Name \ Date"
    ws["A2"].font = Font(bold=True)
    while run:
        daysAfter = daysAfter + 1
        date = datetime.date.today() + datetime.timedelta(days=daysAfter)
        if(date.month != thisMonth and date.month != nextMonth and daysAfter > 0):
            run = False
            break
        if (date.month == nextMonth and month == 'next')\
            or (date.month == prevMonth and month == 'last')\
            or (date.month == thisMonth and month == 'this'):
            col = col + 1
            char = get_column_letter(col)
            ws[char + "1"] = date
            ws[char + "1"].font = Font(bold=True, italic=True, size=9)
            ws[char + '1'].border = border
            weekDay = getWeekDay(date.weekday())
            ws[char + "2"] = weekDay
            ws[char + "2"].font = Font(bold=True)
            ws[char + '2'].border = border
            if weekDay == 'Saturday' or weekDay == 'Sunday':
                ws[char + "2"].fill = weekendColor
            else:
                ws[char + "2"].fill = dayColor

    staffPath = "excel/people.xlsx"
    myStaff = peopleManager.Staff(staffPath)
    people = myStaff.people

    row = 3
    columns = col + 1
    for person in people:
        ws.append([person["Name"]])
        ws['A' + str(row)].font = Font(bold=True)
        for col in range(1, columns):
            char = get_column_letter(col)
            if(row % 2) == 0:
                ws[char + str(row)].fill = color1
            else:
                ws[char + str(row)].fill = color2
            ws[char + str(row)].border = border
        row = row + 1

    return wb
