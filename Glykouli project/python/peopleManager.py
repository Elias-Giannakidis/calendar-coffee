from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import Person
from pathManager import getNextMonthPath, getThisMonthPath, getPreviousMonthPath
from calendarManager import getCalendarDays
from openpyxl.styles import Font

path = "excel/people.xlsx"

class Staff:
    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(path)
        self.ws = self.wb.active
        self.countPeople()
        self.getConsts()
        self.getPersons()

    def countPeople(self):
        self.names = []
        self.peopleNum = 0
        for i in range(2, 10):
            name = self.ws['A' + str(i)].value
            if (name == None):
                break
            self.peopleNum = i - 1
            self.names.append(name)

    def getConsts(self):
        self.consts = []
        for col in range(1, 30):
            char = get_column_letter(col)
            const = self.ws[char + '1'].value
            if(const == None):
                break
            self.consts.append(const)


    def getPersons(self):
        self.people = []
        for i in range(2, self.peopleNum + 2):
            col = 1
            person = {}
            for const in self.consts:
                char = get_column_letter(col)
                value = self.ws[char + str(i)].value
                col = col + 1
                if(value != None):
                    person[const] = value
            self.people.append(person)

def generatePeople(monthPath):

    days = getCalendarDays(monthPath)

    # Count how many days the month has
    monthDays = 0
    for day in days:
        monthDays = monthDays + 1

    staff_path = "excel/people.xlsx"
    myStaff = Staff(path=staff_path)

    # People of the shop
    people = Person.genPeople(myStaff.people)

    # Define the person calendar
    pCount = 0
    for person in people:
        pCount = pCount + 1
        calendar = []
        for day in days:
            if person.name in day:
                calendar.append(day[person.name])
        person.calendar = calendar

    # Define repo-motivo days
    repoWeek = int(7/pCount)
    for person in people:
        person.repoWeek = repoWeek

    return people
