from openpyxl.styles import PatternFill

from Person import Person, genPeople
from peopleManager import generatePeople
from openpyxl import Workbook, load_workbook
from pathManager import getNextMonthPath, getThisMonthPath, getPreviousMonthPath
from calendarManager import getCalendarDays, getCalendarIsValid, getCalendarScore, getVardiesList
from openpyxl.utils import get_column_letter
from excelManager import genWorkSheet

import itertools
import random
BUFFER_SIZE = 20
FREQ = 7

print('\nΛοιπόν γλυκούλι άκου τι θα πατίσεις <3')
print('---------')
print('Για τον προιγούμενο μήνα πατάς: "last"')
print('Για αυτόν τον μήνα πατάς: "this"')
print('Και για τον επόμενο μήνα πατάς: "next"')
print('---------------')
month = input('Άρα γλυκούλι... Για ποιο μήνα θες να σου φτιάξω το excel-ακι? :) \n')

def getMonthPath(month):
    if month == 'next':
        path = getNextMonthPath()
    elif month == 'this':
        path = getThisMonthPath()
    elif month == 'last':
        path = getPreviousMonthPath()
    else:
        month = input('Ρε γλυκούλι κάτι δε πάτησες καλά. Πάμε ξανά! \n')
        path = getMonthPath(month)
    return path

monthPath = getMonthPath(month)
people = generatePeople(monthPath)
vardiesList = getVardiesList()

# print people!
# for person in people:
#     person.printMe()


days = getCalendarDays(monthPath)
monthDays = 0
for calendar in days:
    monthDays = monthDays + 1

# Loop to make the possible calendars
calendars = [[]]
count = 0
for intDay in range(monthDays):
    count = count + 1
    newCalendars = []
    for vardies in vardiesList:  # vardies = [ morning, evening, repo ]
        for calendar in calendars:   # calendar = array(vardies)
            newCalendar = calendar.copy()
            newCalendar.append(vardies)
            if getCalendarIsValid(people, newCalendar):
                newCalendars.append(newCalendar)
    if count % FREQ == 0:
        newCalendars = sorted(newCalendars, key=lambda calendar: -1 * getCalendarScore(people, calendar))
        newCalendars = newCalendars[: BUFFER_SIZE]
    calendars = newCalendars
    print('Υπολογίσα την μέρα ', intDay + 1, ' :)')
calendars = sorted(calendars, key=lambda calendar: -1 * getCalendarScore(people, calendar))
calendars = calendars[: BUFFER_SIZE]
print('\nΤα υπολόγισα όλλα και σου φτιάχνω το excel-ακι :)\n')

def makeExcel(index):
    path = monthPath
    program = calendars[index]
    wb = load_workbook(path)
    sheet = 'Schedule_' + str(index)
    wb = genWorkSheet(wb, sheet, month)
    ws = wb[sheet]
    # make the first column
    for col in range(2, monthDays + 2):
        char = get_column_letter(col)
        dayProgramm = program[col - 2]
        i = 3
        for personDayProgram in dayProgramm:
            ws[char + str(i)] = personDayProgram
            personScedule = people[i - 3].calendar[col - 2]['value']
            if(personDayProgram == 'repo'):
                ws[char + str(i)].fill = PatternFill(fill_type='solid', start_color="78E53F")
            if(personDayProgram != personScedule and personScedule != ''):
                ws[char + str(i)].fill = PatternFill(fill_type='solid', start_color="ff0000")
            i = i + 1
    wb.save(path)


for i in range(1, BUFFER_SIZE):
    # print('The calendar ', i, 'has score ', getCalendarScore(people, calendars[i]))
    makeExcel(i)

print('\n\n')
print('  ***   ***')
print(' *    *    *')
print('* Στεφανία! *')
print(' *         *')
print('   *     *')
print('     * *')
print('      | ')
print('     /')
print('     \ ')
print('      \___ Το Excel-άκι σου είναι έτοιμο :)')


