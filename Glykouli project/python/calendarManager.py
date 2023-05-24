
import itertools
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def getCalendarDays(path):
    wb = load_workbook(path)
    ws = wb.active
    days = []
    for col in range(2, 35):
        day = {}
        for row in range(3, 10):
            char = get_column_letter(col)
            name = ws['A' + str(row)].value
            if(name != None):
                weekday = ws[char + '2'].value
                date = ws[char + '1'].value
                if(weekday != None):
                    value = ws[char + str(row)].value
                    if(value!= None):
                        day[name] = {
                            'name': name,
                            'weekday': weekday,
                            'date': date,
                            'value': value
                        }
                    else:
                        day[name] = {
                            'name': name,
                            'weekday': weekday,
                            'date': date,
                            'value': ''
                        }
        if(day != {}):
             days.append(day)
    return days

def getPossibelVardies(path):
    wb = load_workbook(path)
    ws = wb.active
    days = []
    col = 0
    while True:
        col = col + 1
        char = get_column_letter(col)
        run = True
        day = []
        row = 0
        while run:
            row = row + 1
            vardia = ws[char + str(row)].value
            if(vardia == None):
                run = False
            else:
                day.append(vardia)
        if(day == []):
            break
        days.append(day)
    return days

def getVardiesList():
    vardies_path = 'excel/vardies.xlsx'
    # An array with arrays with the possible work days - vardies
    vardiesList = getPossibelVardies(vardies_path)
    newVardiesList = []
    for vardies in vardiesList:
        posVardiesList = itertools.permutations(vardies)
        for posVardies in posVardiesList:
            newVardiesList.append(posVardies)
    vardiesList = newVardiesList
    vardiesList = list(set(vardiesList))
    return vardiesList

def getCalendarScore(people, calendar):
    j = 0
    score = 0
    minRepoScore = 0
    for person in people:
        newCalendar = []
        for day in calendar:
            newCalendar.append(day[j])
        j = j + 1
        score = score + person.getScore(newCalendar)
        score = score + person.getRepoScore(calendar)
        minRepoScore = min(score, person.getRepoScore(calendar))
    score = score + minRepoScore
    return score

def getCalendarIsValid(people, calendar):
    valid = True
    j = 0
    for person in people:
        newCalendar = []
        for day in calendar:
            newCalendar.append(day[j])
        j = j + 1
        valid = valid and person.getCalendarValid(newCalendar)
    return valid

