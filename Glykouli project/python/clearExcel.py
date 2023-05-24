from openpyxl.styles import PatternFill

from Person import Person, genPeople
from peopleManager import generatePeople
from openpyxl import Workbook, load_workbook
from pathManager import getNextMonthPath, getThisMonthPath, getPreviousMonthPath
from calendarManager import getCalendarDays, getCalendarIsValid, getCalendarScore, getVardiesList
from openpyxl.utils import get_column_letter
from excelManager import genWorkSheet

print('\nΛοιπόν γλυκούλι άκου τι θα πατίσεις <3')
print(' Είναι τα ίδια όπως όταν έφτιαχνες το excel!')
print('Το Main δε στο πηράζει μη φοβάσε. Ότι έχεις γράψει εκεί θα μίνει ίδιο <3')
print('---------')
print('Για τον προιγούμενο μήνα πατάς: "last"')
print('Για αυτόν τον μήνα πατάς: "this"')
print('Και για τον επόμενο μήνα πατάς: "next"')
print('---------------')
month = input('Άρα γλυκούλι... Για ποιο μήνα θες να σου καθαρίσω το excel-ακι? :) \n')

def getPath(month):
    if month == 'last':
        path = getPreviousMonthPath()
        print('Τέλεια σου καθαρίζω το excel για τον προηγούμενο μήνα ;)')
    elif month == 'this':
        path = getThisMonthPath()
        print('Τέλεια σου καθαρίζω το excel για αυτόν τον μήνα ;)')
    elif month == 'next':
        path = getNextMonthPath()
        print('Τέλεια σου καθαρίζω το excel για τον επόμενο μήνα')
    else:
        month = input('Ρε γλυκούλι κάτι δε πάτησες καλά ξανά πάμε \n')
        path = getPath(month)
    return path

path = getPath(month)

wb = load_workbook(path)

for i in range(1, 20):
    sheet = 'Schedule_' + str(i)
    ws = wb[sheet]
    wb.remove(ws)
    wb.save(path)

print('\n\n\n')
print('  ***   ***')
print(' *    *    *')
print('* Στεφανία! *')
print(' *         *')
print('   *     *')
print('     * *')
print('      | ')
print('     /')
print('     \ ')
print('      \___ Το Excel-άκι σου είναι καθαρό! :)')